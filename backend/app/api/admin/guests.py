"""Admin guest management endpoints."""

import csv
import io
import json
import logging
from datetime import datetime
from uuid import UUID, uuid4

import openpyxl
from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy import func, select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.deps import get_current_user
from app.database import get_session
from app.models import User
from app.models.role import RoleCode
from app.models.user import GuestSubtype
from app.models.user_role import UserRole
from app.schemas.admin import (
    GuestDetailResponse,
    GuestListItem,
    GuestUploadResult,
)
from app.schemas.expert import RowError
from app.schemas.merge import MergeApplyResult
from app.services.admin import admin_service, audit_service, dedup_service, merge_service
from app.services.core import user_service

logger = logging.getLogger(__name__)

router = APIRouter()


@router.get("/guests", response_model=list[GuestListItem])
async def list_guests(
    search: str | None = None,
    subtype: str | None = None,
    role: str | None = None,
    source: str | None = None,
    db: AsyncSession = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """List all guests and business partners with profile summaries."""
    event = await user_service.get_current_event(db)
    if not event:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="No active event"
        )

    return await admin_service.list_guests(db, event.id, search, subtype, role, source=source)


@router.get("/guests/export")
async def export_guests(
    search: str | None = None,
    subtype: str | None = None,
    role: str | None = None,
    source: str | None = None,
    db: AsyncSession = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Export guests to Excel file."""
    event = await user_service.get_current_event(db)
    if not event:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="No active event"
        )

    guests = await admin_service.list_guests(db, event.id, search, subtype, role, source=source)

    # Create Excel workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Гости"

    # Header
    ws.append([
        "ФИО",
        "Телеграм",
        "Роль",
        "Подтип",
        "Интересы",
        "Цели",
        "Резюме профиля",
        "Компания",
        "Должность",
        "Бизнес-цели",
        "Статус партнёра",
        "Рекомендаций",
        "Запросов контактов",
    ])

    # Data rows
    for guest in guests:
        # GuestListItem doesn't have interests/goals/etc - only tags and keywords
        ws.append([
            guest.full_name,
            f"@{guest.username}" if guest.username else "",
            guest.role,
            guest.guest_subtype or "",
            ", ".join(guest.tags) if guest.tags else "",
            ", ".join(guest.keywords) if guest.keywords else "",
            guest.profile_summary or "",
            "",  # company - not in GuestListItem
            "",  # position - not in GuestListItem
            "",  # business_objectives - not in GuestListItem
            "",  # partner_status - not in GuestListItem
            guest.recommendations_count or 0,
            guest.contact_requests_count or 0,
        ])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Generate filename with timestamp
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"guests_{event.name.replace(' ', '_')}_{timestamp}.xlsx"

    await audit_service.log_action(
        db, current_user, "export_guests",
        entity_type="guests",
        details={"count": len(guests), "filters": {"search": search, "subtype": subtype, "role": role}},
    )

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.delete("/guests/all")
async def delete_all_guests(
    subtype: str | None = Query(None),
    db: AsyncSession = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Delete all guests for current event, optionally filtered by subtype."""
    event = await user_service.get_current_event(db)
    if not event:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No active event")

    guest_role = await user_service.get_role_by_code(db, RoleCode.GUEST)
    if not guest_role:
        raise HTTPException(status_code=500, detail="Guest role not found")

    query = select(UserRole).join(User, UserRole.user_id == User.id).where(
        UserRole.event_id == event.id,
        UserRole.role_id == guest_role.id,
    )
    if subtype:
        try:
            subtype_enum = GuestSubtype(subtype)
        except ValueError:
            valid = [s.value for s in GuestSubtype]
            raise HTTPException(status_code=422, detail=f"Invalid subtype. Valid: {valid}")
        query = query.where(User.guest_subtype == subtype_enum)

    result = await db.execute(query)
    roles = list(result.scalars().all())
    count = len(roles)

    user_ids = [ur.user_id for ur in roles]
    for ur in roles:
        await db.delete(ur)

    # Delete users with synthetic IDs (created by import)
    for uid in user_ids:
        user = await db.get(User, uid)
        if user and user.telegram_user_id.startswith("guest-"):
            await db.delete(user)

    await db.commit()

    await audit_service.log_action(
        db, current_user, "delete_all_guests",
        entity_type="guests",
        details={"deleted": count, "subtype": subtype},
    )

    return {"deleted": count}


@router.get("/guests/{user_id}", response_model=GuestDetailResponse)
async def get_guest_detail(
    user_id: UUID,
    db: AsyncSession = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Get detailed guest profile."""
    event = await user_service.get_current_event(db)
    if not event:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="No active event"
        )

    try:
        return await admin_service.get_guest_detail(db, event.id, user_id)
    except ValueError as e:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail=str(e))


@router.post("/guests/upload")
async def upload_guests(
    file: UploadFile = File(...),
    default_subtype: str = Query(...),
    confirm_replace: bool = Query(False),
    db: AsyncSession = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Bulk import guests from CSV or JSON file."""

    event = await user_service.get_current_event(db)
    if not event:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND, detail="No active event"
        )

    # Validate subtype
    try:
        subtype_enum = GuestSubtype(default_subtype)
    except ValueError:
        valid = [s.value for s in GuestSubtype]
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Invalid subtype '{default_subtype}'. Valid: {valid}",
        )

    # Get guest role
    guest_role = await user_service.get_role_by_code(db, RoleCode.GUEST)
    if not guest_role:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail="Guest role not found in database",
        )

    # Check existing guest users for this event
    existing_count = await db.scalar(
        select(func.count(UserRole.id)).where(
            UserRole.event_id == event.id,
            UserRole.role_id == guest_role.id,
        )
    )

    if existing_count and existing_count > 0 and not confirm_replace:
        return GuestUploadResult(
            total_parsed=0,
            imported=0,
            duplicates=0,
            errors=[],
        ).model_dump() | {
            "existing_count": existing_count,
            "message": f"{existing_count} guests already exist. Set confirm_replace=true to replace.",
        }

    if existing_count and existing_count > 0 and confirm_replace:
        # Delete existing guest user_roles and their users for this event
        guest_user_roles = (
            await db.execute(
                select(UserRole).where(
                    UserRole.event_id == event.id,
                    UserRole.role_id == guest_role.id,
                )
            )
        ).scalars().all()
        guest_user_ids = [ur.user_id for ur in guest_user_roles]
        for ur in guest_user_roles:
            await db.delete(ur)
        # Delete users that were created for import (synthetic telegram_user_id)
        for uid in guest_user_ids:
            user = await db.get(User, uid)
            if user and user.telegram_user_id.startswith("guest-"):
                await db.delete(user)
        await db.flush()

    # Parse file
    content = await file.read()
    file_hash = dedup_service.compute_file_hash(content)
    dup_info = await dedup_service.check_recent_duplicate(db, file_hash, "upload_guests")
    filename = file.filename or ""
    rows: list[dict] = []

    if filename.endswith(".json"):
        try:
            data = json.loads(content.decode("utf-8"))
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid JSON file")
        if not isinstance(data, list):
            raise HTTPException(status_code=400, detail="Expected JSON array")
        rows = data
    elif filename.endswith(".csv"):
        try:
            text = content.decode("utf-8")
            reader = csv.DictReader(io.StringIO(text))
            rows = list(reader)
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid CSV file")
    elif filename.endswith(".xlsx"):
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            rows_iter = ws.iter_rows(values_only=True)
            headers = [str(h).strip() if h else f"col_{i}" for i, h in enumerate(next(rows_iter))]
            for row_values in rows_iter:
                if all(v is None for v in row_values):
                    continue
                rows.append({headers[i]: (str(v).strip() if v is not None else "") for i, v in enumerate(row_values)})
            wb.close()
        except Exception:
            raise HTTPException(status_code=400, detail="Invalid XLSX file")
    else:
        raise HTTPException(status_code=400, detail="Unsupported file format. Use .csv, .json or .xlsx")

    # Process rows
    imported = 0
    duplicates = 0
    errors: list[RowError] = []
    seen_names: set[str] = set()

    for i, row in enumerate(rows):
        # Normalize keys to lowercase for case-insensitive matching
        row_lower = {k.lower().strip(): v for k, v in row.items()}

        # Helper to get value with fallback to lowercase keys
        def get_field(*keys):
            for k in keys:
                val = row.get(k) or row_lower.get(k.lower())
                if val:
                    return val
            return ""

        name = get_field("name", "фио *", "фио", "имя *", "имя").strip()
        if not name:
            errors.append(RowError(row=i + 1, field="name", message="Имя обязательно"))
            continue

        # Check duplicate within file
        name_lower_key = name.lower()
        if name_lower_key in seen_names:
            duplicates += 1
            continue
        seen_names.add(name_lower_key)

        telegram = get_field("telegram", "телеграм *", "телеграм").strip()
        username = telegram.lstrip("@") if telegram else None

        # Generate synthetic telegram_user_id
        synthetic_id = f"guest-{uuid4().hex[:8]}"

        user = await user_service.upsert_user(
            db,
            telegram_user_id=synthetic_id,
            full_name=name,
            username=username,
            source="import",
        )

        await user_service.set_role(
            db,
            user_id=user.id,
            event_id=event.id,
            role=guest_role,
            guest_subtype=subtype_enum,
        )

        imported += 1

    await audit_service.log_action(
        db, current_user, "upload_guests",
        entity_type="guests",
        details={"imported": imported, "duplicates": duplicates, "errors": len(errors), "file_hash": file_hash},
    )

    return GuestUploadResult(
        total_parsed=len(rows),
        imported=imported,
        duplicates=duplicates,
        errors=errors,
        duplicate_warning=dup_info["warning"] if dup_info else None,
    )


@router.post("/guests/upload/preview")
async def preview_guest_upload(
    file: UploadFile = File(...),
    default_subtype: str = Query(...),
    db: AsyncSession = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Dry-run analysis of guest file vs DB. Returns MergePreview."""
    event = await user_service.get_current_event(db)
    if not event:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No active event")

    try:
        subtype_enum = GuestSubtype(default_subtype)
    except ValueError:
        valid = [s.value for s in GuestSubtype]
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Invalid subtype '{default_subtype}'. Valid: {valid}",
        )

    content = await file.read()
    filename = file.filename or "file.xlsx"

    try:
        preview, _ = await merge_service.analyze_guest_merge(
            db, event.id, content, filename, subtype_enum,
        )
    except (ValueError, IndexError, KeyError) as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.exception("Guest preview failed: %s", e)
        raise HTTPException(status_code=500, detail=f"Ошибка анализа файла: {e}")

    return preview


@router.post("/guests/upload/merge", response_model=MergeApplyResult)
async def merge_guest_upload(
    file: UploadFile = File(...),
    default_subtype: str = Query(...),
    add_new: bool = Query(True),
    update_existing: bool = Query(True),
    db: AsyncSession = Depends(get_session),
    current_user: User = Depends(get_current_user),
):
    """Smart merge: add new + update changed guests."""
    event = await user_service.get_current_event(db)
    if not event:
        raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="No active event")

    try:
        subtype_enum = GuestSubtype(default_subtype)
    except ValueError:
        valid = [s.value for s in GuestSubtype]
        raise HTTPException(
            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
            detail=f"Invalid subtype '{default_subtype}'. Valid: {valid}",
        )

    content = await file.read()
    filename = file.filename or "file.xlsx"

    try:
        _, internal = await merge_service.analyze_guest_merge(
            db, event.id, content, filename, subtype_enum,
        )
    except (ValueError, IndexError, KeyError) as e:
        raise HTTPException(status_code=400, detail=str(e))
    except Exception as e:
        logger.exception("Guest merge analyze failed: %s", e)
        raise HTTPException(status_code=500, detail=f"Ошибка анализа файла: {e}")

    result = await merge_service.apply_guest_merge(
        db, event.id, subtype_enum, internal,
        add_new=add_new, update_existing=update_existing,
    )

    await audit_service.log_action(
        db, current_user, "merge_guests",
        entity_type="guests",
        details={
            "added": result.added, "updated": result.updated,
            "skipped": result.skipped, "subtype": default_subtype,
        },
    )

    return result
