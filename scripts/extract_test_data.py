"""
Extract structured test datasets from anonymized Excel files.

Produces:
  data/test/projects.json    — 330 projects with scores, rooms, tags
  data/test/experts.json     — experts with profiles + linked IDs
  data/test/students.json    — student profiles from checkpoints

Usage: python scripts/extract_test_data.py
"""

import json
import re
import sys
from pathlib import Path
from collections import defaultdict
import openpyxl

REPO = Path(r"C:\Users\Admin\Desktop\AI Talent Camp\demoday-core")
TEST_DIR = REPO / "data" / "test"

# --- 1. EXPERTS ---

def extract_experts():
    """Merge experts-public.json with anonymized Excel IDs."""

    # Load public expert profiles
    with open(REPO / "data" / "experts-public.json", "r", encoding="utf-8") as f:
        public = json.load(f)
    exp_profiles = {e["id"]: e for e in public}

    # Load private mapping for bridge
    with open(REPO / "data" / "expert-mapping.json", "r", encoding="utf-8") as f:
        priv = json.load(f)
    name_to_exp = {e["name"].strip(): e["id"] for e in priv}

    # Load anonymization mapping
    with open(TEST_DIR / "_anonymization_mapping.json", "r", encoding="utf-8") as f:
        anon_map = json.load(f)
    expert_anon = anon_map.get("experts", {})

    # Build bridge: Эксперт_NNN -> EXP-NNN
    bridge = {}
    for real_name, anon_id in expert_anon.items():
        if real_name in name_to_exp:
            bridge[anon_id] = name_to_exp[real_name]
        else:
            parts = real_name.split()
            if len(parts) >= 2:
                rev = " ".join(parts[1:]) + " " + parts[0]
                if rev in name_to_exp:
                    bridge[anon_id] = name_to_exp[rev]
                    continue
            for pname, pid in name_to_exp.items():
                if parts[0] in pname and (len(parts) < 2 or parts[-1] in pname):
                    bridge[anon_id] = pid
                    break

    # Build merged dataset
    experts = []
    for anon_id in sorted(expert_anon.values(), key=lambda x: int(re.search(r'\d+', x).group())):
        entry = {"anon_id": anon_id}
        exp_id = bridge.get(anon_id)
        if exp_id and exp_id in exp_profiles:
            profile = exp_profiles[exp_id]
            entry["profile_id"] = exp_id
            entry["position"] = profile.get("position", "")
            entry["expertise_tags"] = profile.get("expertise_tags", [])
            entry["dd_status"] = profile.get("dd_status", "")
            entry["expert_type"] = profile.get("expert_type", "")
        else:
            entry["profile_id"] = None
            entry["position"] = ""
            entry["expertise_tags"] = []
            entry["dd_status"] = ""
            entry["expert_type"] = ""
        experts.append(entry)

    out = TEST_DIR / "experts.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(experts, f, ensure_ascii=False, indent=2)

    linked = sum(1 for e in experts if e["profile_id"])
    print(f"  Experts: {len(experts)} total, {linked} linked to profiles -> {out.name}")
    return bridge


def _is_time(val):
    return bool(re.match(r"^\d{1,2}:\d{2}", str(val)))


# --- 2. PROJECTS ---

def extract_projects(expert_bridge):
    """Extract projects from anonymized expert evaluation files."""

    projects = {}  # title -> project dict

    for tag, fn in [
        ("22jan", TEST_DIR / "experts_22jan_anon.xlsx"),
        ("23jan", TEST_DIR / "experts_23jan_anon.xlsx"),
    ]:
        wb = openpyxl.load_workbook(fn, read_only=True)

        for sn in wb.sheetnames:
            ws = wb[sn]
            rows = list(ws.iter_rows(values_only=True))
            if not rows or len(rows) < 2:
                continue

            # Get room/topic from row 0
            room = str(rows[0][1]).strip() if len(rows[0]) > 1 and rows[0][1] else sn

            # Determine format
            sn_lower = sn.lower()
            if "бизнес" in sn_lower:
                fmt = "business"
            elif "прожарка" in sn_lower:
                fmt = "roast"
            elif "aitalconf" in sn_lower or "conf" in sn_lower:
                fmt = "research"
            else:
                fmt = "demo_day"

            day = "2026-01-22" if tag == "22jan" else "2026-01-23"

            # Parse blocks
            for i, r in enumerate(rows):
                if not r or not r[0]:
                    continue
                if "ФИО" not in str(r[0]):
                    continue
                if i + 1 >= len(rows):
                    continue

                student_row = rows[i + 1]
                # Determine which col has the project title
                col_a = str(student_row[0]).strip() if student_row[0] else ""
                col_b = str(student_row[1]).strip() if len(student_row) > 1 and student_row[1] else ""

                # Find time from nearby rows
                time_slot = ""
                for k in range(max(0, i - 2), i):
                    if rows[k] and rows[k][0] and _is_time(rows[k][0]):
                        time_slot = str(rows[k][0])
                        if ":" in time_slot and len(time_slot) > 8:
                            # datetime.time object rendered as HH:MM:SS
                            time_slot = time_slot[:5]
                        break

                # Project title is usually in col B, student in col A
                # But sometimes swapped
                title = col_b.replace("\n", " ").strip() if col_b else col_a.replace("\n", " ").strip()
                if not title or title == "None":
                    continue

                # Clean title
                title = re.sub(r'\s+', ' ', title).strip().rstrip('"').strip()

                # Collect expert scores
                evaluations = []
                j = i + 2
                while j < len(rows):
                    er = rows[j]
                    if not er or not er[0]:
                        break
                    ename = str(er[0]).strip()
                    if not ename or ename == "None":
                        break
                    if "ФИО" in ename or _is_time(ename) or "Артефакты" in ename:
                        break

                    scores = {}
                    criteria_names = ["актуальность", "практ_значимость", "новизна",
                                     "импакт", "rd_сложность", "масштабирование", "критерий_7"]
                    for k in range(2, min(9, len(er))):
                        if er[k]:
                            try:
                                s = float(er[k])
                                if 0 < s <= 3:
                                    scores[criteria_names[k - 2]] = s
                            except (ValueError, TypeError):
                                pass

                    comment = ""
                    if len(er) > 10 and er[10]:
                        comment = str(er[10]).strip()[:500]

                    if scores:
                        evaluations.append({
                            "expert_anon_id": ename,
                            "scores": scores,
                            "comment": comment if comment and comment != "None" else "",
                        })
                    j += 1

                # Aggregate into project
                if title not in projects:
                    projects[title] = {
                        "title": title,
                        "rooms": [],
                        "days": [],
                        "formats": [],
                        "time_slots": [],
                        "evaluations": [],
                        "tags": [],
                    }

                p = projects[title]
                if room not in p["rooms"]:
                    p["rooms"].append(room)
                if day not in p["days"]:
                    p["days"].append(day)
                if fmt not in p["formats"]:
                    p["formats"].append(fmt)
                if time_slot and time_slot not in p["time_slots"]:
                    p["time_slots"].append(time_slot)
                p["evaluations"].extend(evaluations)

        wb.close()

    # Compute aggregated scores per project
    for p in projects.values():
        if p["evaluations"]:
            all_scores = []
            for ev in p["evaluations"]:
                all_scores.extend(ev["scores"].values())
            p["avg_score"] = round(sum(all_scores) / len(all_scores), 2) if all_scores else 0
            p["expert_count"] = len(p["evaluations"])
        else:
            p["avg_score"] = 0
            p["expert_count"] = 0

    # Infer tags from room names
    tag_map = {
        "EdTech": ["EdTech", "Образование", "Education"],
        "NLP": ["NLP", "Языки"],
        "CV": ["CV"],
        "LLM": ["LLM"],
        "FinTech": ["FinTech"],
        "RecSys": ["RecSys", "HR"],
        "Security": ["Security", "Безопасность"],
        "Audio": ["Audio", "Распознавание речи", "речи"],
        "ML": ["ML в промышленности", "Промышленность"],
        "Agents": ["Agent", "агент", "Мультиагентные"],
        "RL": ["RL"],
        "BioTech": ["BioTech", "Science"],
        "MedTech": ["MedTech", "Медицина"],
        "Business": ["Бизнес", "Продажи", "B2B", "маркетинг"],
        "Geo": ["Гео", "Туризм"],
        "3D/Games": ["Геймификация", "3D", "визуализация"],
    }

    for p in projects.values():
        tags = set()
        for room in p["rooms"]:
            for tag_name, keywords in tag_map.items():
                if any(kw.lower() in room.lower() for kw in keywords):
                    tags.add(tag_name)
        p["tags"] = sorted(tags)

    # Sort by title and output
    result = sorted(projects.values(), key=lambda x: x["title"])

    out = TEST_DIR / "projects.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    scored = sum(1 for p in result if p["avg_score"] > 0)
    print(f"  Projects: {len(result)} total, {scored} with scores -> {out.name}")


# --- 3. STUDENTS ---

def extract_students():
    """Extract student profiles from checkpoint files."""

    students = {}  # anon_id -> profile

    # Checkpoint 3 — richest data
    fn3 = TEST_DIR / "checkpoint3_anon.xlsx"
    wb = openpyxl.load_workbook(fn3, read_only=True)
    ws = wb["Финальная выгрузка"]
    rows = list(ws.iter_rows(values_only=True))
    headers = rows[0]

    tech_cols = {
        48: "CV", 49: "NLP", 50: "LLM", 51: "Agentic_AI",
        52: "VLM", 53: "ML", 54: "MLOps", 55: "RecSys",
        56: "Audio", 57: "Data_Scientist", 58: "Other",
    }

    for row in rows[1:]:
        if not row or not row[3]:
            continue

        # Col 3 = student being reviewed (always present)
        anon_id = str(row[3]).strip()
        if not anon_id or "Студент_" not in anon_id:
            continue

        review_format = str(row[2]).strip() if row[2] else ""

        # Tech areas only available in Self-review
        tech_areas = []
        if review_format == "Self-review":
            for col_idx, tech_name in tech_cols.items():
                if col_idx < len(row) and row[col_idx]:
                    tech_areas.append(tech_name)

        track = str(row[8]).strip() if row[8] else ""
        course = str(row[7]).strip() if row[7] else ""
        project = str(row[9]).strip() if row[9] else ""

        # Mentor involvement
        mentor_score = None
        if row[12]:
            try:
                mentor_score = float(row[12])
            except (ValueError, TypeError):
                pass

        # Goal achievement (Self-review only)
        goal_score = None
        if review_format == "Self-review" and len(row) > 17 and row[17]:
            try:
                goal_score = float(row[17])
            except (ValueError, TypeError):
                pass

        if anon_id not in students:
            students[anon_id] = {
                "anon_id": anon_id,
                "track": track if track and track != "N/A" else "",
                "course": course if course and course != "N/A" else "",
                "project": project if project and project != "None" else "",
                "tech_areas": tech_areas,
                "mentor_involvement": mentor_score,
                "goal_achievement": goal_score,
                "review_formats": [],
            }
        s = students[anon_id]
        if review_format and review_format not in s["review_formats"]:
            s["review_formats"].append(review_format)
        if not s["track"] and track and track != "N/A":
            s["track"] = track
        if not s["course"] and course and course != "N/A":
            s["course"] = course
        if not s["project"] and project and project != "None":
            s["project"] = project
        if tech_areas and not s["tech_areas"]:
            s["tech_areas"] = tech_areas
        if mentor_score and not s["mentor_involvement"]:
            s["mentor_involvement"] = mentor_score
        if goal_score and not s["goal_achievement"]:
            s["goal_achievement"] = goal_score

    wb.close()

    # Enrich from checkpoint 1&2 (more students, track info)
    fn12 = TEST_DIR / "checkpoint12_anon.xlsx"
    wb2 = openpyxl.load_workbook(fn12, read_only=True)

    for sn in ["1КР 13.10"]:
        ws2 = wb2[sn]
        rows2 = list(ws2.iter_rows(values_only=True))
        if not rows2:
            continue
        for row in rows2[1:]:
            if not row or len(row) < 7 or not row[2]:
                continue
            anon_id = str(row[2]).strip()
            if "Студент_" not in anon_id:
                continue
            track = str(row[5]).strip() if len(row) > 5 and row[5] else ""
            course = str(row[4]).strip() if len(row) > 4 and row[4] else ""
            project = str(row[6]).strip() if len(row) > 6 and row[6] else ""

            if anon_id not in students:
                students[anon_id] = {
                    "anon_id": anon_id,
                    "track": track if track and track != "N/A" else "",
                    "course": course if course and course != "N/A" else "",
                    "project": project if project and project != "None" else "",
                    "tech_areas": [],
                    "mentor_involvement": None,
                    "goal_achievement": None,
                    "review_formats": [],
                }
            else:
                s = students[anon_id]
                if not s["track"] and track and track != "N/A":
                    s["track"] = track
                if not s["course"] and course and course != "N/A":
                    s["course"] = course
                if not s["project"] and project and project != "None":
                    s["project"] = project

    wb2.close()

    result = sorted(students.values(), key=lambda x: int(re.search(r'\d+', x["anon_id"]).group()))

    out = TEST_DIR / "students.json"
    with open(out, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    with_track = sum(1 for s in result if s["track"])
    with_tech = sum(1 for s in result if s["tech_areas"])
    with_project = sum(1 for s in result if s["project"])
    print(f"  Students: {len(result)} total, {with_track} with track, {with_tech} with tech areas, {with_project} with project -> {out.name}")


def main():
    print("Extracting test datasets...\n")

    print("1. Experts:")
    bridge = extract_experts()

    print("\n2. Projects:")
    extract_projects(bridge)

    print("\n3. Students:")
    extract_students()

    print("\nDone.")


if __name__ == "__main__":
    main()
