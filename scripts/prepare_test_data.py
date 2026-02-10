#!/usr/bin/env python3
"""
Prepare test data files from checkpoint12_anon.xlsx in the correct format for import.
"""
import openpyxl
import json
import re
from pathlib import Path

# Paths
DATA_DIR = Path(__file__).parent.parent / "data" / "test"
OUTPUT_DIR = Path(__file__).parent.parent / "data" / "import_ready"
OUTPUT_DIR.mkdir(exist_ok=True)

# Tag normalization: Russian/long names -> short English
TAG_NORMALIZE = {
    'EdTech': 'EdTech',
    'NLP': 'NLP',
    'CV': 'CV',
    'Computer Vision': 'CV',
    'ML': 'ML',
    'Machine Learning': 'ML',
    'DL': 'DL',
    'Deep Learning': 'DL',
    'LLM': 'LLM',
    'RAG': 'RAG',
    'Автономные агенты': 'Agents',
    'Агентные системы': 'Agents',
    'Agents': 'Agents',
    'Recsys': 'Recsys',
    'предиктивная аналитика': 'Recsys',
    'FinTech': 'FinTech',
    'финтех': 'FinTech',
    'ML в Fintech': 'FinTech',
    'MedTech': 'MedTech',
    'AgriTech': 'AgriTech',
    'агротех': 'AgriTech',
    'Security': 'Security',
    'MLOps': 'MLOps',
    'ASR': 'ASR',
    'Распознавание речи': 'ASR',
    'Speech Recognition': 'ASR',
    'TTS': 'TTS',
    'Синтез речи': 'TTS',
    'RL': 'RL',
    'Обучение с подкреплением': 'RL',
    'TimeSeries': 'TimeSeries',
    'Временные ряды': 'TimeSeries',
    'Backend': 'Backend',
    'ML в промышленности': 'Industrial',
    'DL в промышленности': 'Industrial',
}

# Load project tags from schedule if available
project_tags_map = {}
tags_file = Path('/tmp/project_tags.json')
if tags_file.exists():
    with open(tags_file, 'r', encoding='utf-8') as f:
        raw_map = json.load(f)
        # Normalize tags
        for project_title, tags_list in raw_map.items():
            normalized_tags = []
            for tag in tags_list:
                normalized = TAG_NORMALIZE.get(tag.strip())
                if normalized and normalized not in normalized_tags:
                    normalized_tags.append(normalized)
            if normalized_tags:
                project_tags_map[project_title] = ', '.join(normalized_tags)
    print(f"Loaded tags for {len(project_tags_map)} projects from schedule")

# Read checkpoint12
checkpoint_file = DATA_DIR / "checkpoint12_anon.xlsx"
wb_checkpoint = openpyxl.load_workbook(checkpoint_file)
ws_checkpoint = wb_checkpoint.active

# Extract projects data
projects = []
students_set = set()

for row_idx in range(2, ws_checkpoint.max_row + 1):
    row = ws_checkpoint[row_idx]

    # Columns: ID, Время, ФИО, Телеграм, Курс, Трек, Название, ...
    fio = row[2].value
    telegram = row[3].value
    track = row[5].value
    title = row[6].value
    plan = row[12].value
    team = row[14].value

    if not title or not fio:
        continue

    # Clean telegram
    if telegram:
        telegram = str(telegram).strip()
        if not telegram.startswith('@'):
            telegram = f'@{telegram}'

    # Use plan as description, limit to 200 chars
    description = str(plan)[:200] if plan else f"Проект {title}"

    # Normalize track to Russian canonical names
    track_map = {
        "Education": "Образовательный",
        "EdTech": "Образовательный",
        "Startup": "Стартап",
        "Industrial": "Индустриальный",
        "Research": "Научный",
    }
    track_normalized = track_map.get(track, None) if track else None

    # Try to match tags from schedule by project title
    tags = ""
    if project_tags_map:
        # Normalize title for matching
        title_normalized = re.sub(r'\s+', ' ', title.strip())
        tags = project_tags_map.get(title_normalized, "")

    projects.append({
        'title': title,
        'description': description,
        'author': fio,
        'telegram_contact': telegram or "",
        'tags': tags,
        'track': track_normalized,
    })

    # Collect students
    if fio and telegram:
        students_set.add((fio, telegram))

    # Add team members to students
    if team:
        members = str(team).split(',')
        for member in members:
            member = member.strip()
            if member and member != fio:
                students_set.add((member, ""))

print(f"Extracted {len(projects)} projects")
print(f"Extracted {len(students_set)} unique students")

# Deduplicate and clean projects
seen_titles = set()
clean_projects = []

for p in projects:
    title = p['title']
    desc = p['description']

    # Skip мусор
    if not title or str(title).strip() in ['-', '']:
        continue
    if not desc or len(str(desc)) < 10:
        continue

    # Skip дубли (case-insensitive)
    title_key = str(title).strip().lower()
    if title_key in seen_titles:
        continue

    seen_titles.add(title_key)
    clean_projects.append(p)

print(f"After dedup/cleanup: {len(clean_projects)} projects")

# Clean students (keep all, names are anonymized as Студент_XXX)
clean_students = students_set

print(f"After cleanup: {len(clean_students)} students")

# Create projects.xlsx
wb_projects = openpyxl.Workbook()
ws_projects = wb_projects.active
ws_projects.title = "Projects"

# Header (Russian)
ws_projects.append(['Название *', 'Описание *', 'Автор *', 'Телеграм', 'Тематики', 'Трек'])

# Data (all clean projects)
for p in clean_projects:
    ws_projects.append([
        p['title'],
        p['description'],
        p['author'],
        p['telegram_contact'],
        p['tags'],
        p['track'],
    ])

projects_file = OUTPUT_DIR / "projects.xlsx"
wb_projects.save(projects_file)
print(f"Created {projects_file}")

# Create students.xlsx
wb_students = openpyxl.Workbook()
ws_students = wb_students.active
ws_students.title = "Students"

# Header (Russian)
ws_students.append(['Имя', 'Телеграм'])

# Data (all clean students)
for name, telegram in sorted(clean_students):
    ws_students.append([name, telegram])

students_file = OUTPUT_DIR / "students.xlsx"
wb_students.save(students_file)
print(f"Created {students_file}")

# Create experts.xlsx (mock data)
wb_experts = openpyxl.Workbook()
ws_experts = wb_experts.active
ws_experts.title = "Experts"

# Header
ws_experts.append(['Имя *', 'Телеграм', 'Описание', 'Тематики', 'Статус', 'Комментарии'])

# Mock experts with relevant tags
experts_data = [
    ('Алексей Иванов', '@expert_ai', 'Senior ML Engineer в TechCorp, 8 лет опыта', 'NLP, CV, ML', 'Подтвердил', 'Готов взять NLP и CV проекты'),
    ('Мария Петрова', '@expert_nlp', 'Head of AI Research в UniversityLab', 'NLP, Research, Education', 'Подтвердил', ''),
    ('Дмитрий Сидоров', '@expert_cv', 'Computer Vision Lead, специализация медицина', 'CV, Deep Learning', 'Подтвердил', 'Интересуют прикладные CV проекты'),
    ('Анна Козлова', '@expert_fintech', 'FinTech Product Manager в банке', 'Finance, Product', 'Уточняется', 'Написала, что постарается приехать'),
    ('Сергей Морозов', '@expert_edu', 'EdTech Founder, бывший преподаватель ВШЭ', 'Education, Startups', 'Подтвердил', ''),
    ('Елена Новикова', '@expert_ml', 'ML Team Lead в e-commerce', 'ML, Data Science', 'Подтвердил', 'Recsys, fintech могу взять'),
    ('Игорь Волков', '@expert_web', 'Senior Frontend Dev, React/Vue эксперт', 'Web, UI/UX', 'Не ответил', ''),
    ('Ольга Смирнова', '@expert_mobile', 'Mobile Dev Lead в стартапе', 'Mobile, iOS, Android', 'Подтвердил', ''),
    ('Павел Федоров', '@expert_backend', 'Backend Architect, highload системы', 'Backend, Databases, API', 'Подтвердил', 'Могу помочь с архитектурой'),
    ('Наталья Егорова', '@expert_ds', 'Data Scientist в консалтинге', 'Data Science, Analytics, ML', 'Отказался', 'Занята в этот день'),
]

for expert in experts_data:
    ws_experts.append(expert)

experts_file = OUTPUT_DIR / "experts.xlsx"
wb_experts.save(experts_file)
print(f"Created {experts_file}")

# Create partners.xlsx (mock data)
wb_partners = openpyxl.Workbook()
ws_partners = wb_partners.active
ws_partners.title = "Partners"

# Header
ws_partners.append(['Имя', 'Телеграм'])

# Mock partners
partners_data = [
    ('Иван Соколов', '@partner_tech'),
    ('Татьяна Лебедева', '@partner_invest'),
    ('Максим Орлов', '@partner_hr'),
    ('Виктория Белова', '@partner_business'),
    ('Андрей Козлов', '@partner_corp'),
]

for partner in partners_data:
    ws_partners.append(partner)

partners_file = OUTPUT_DIR / "partners.xlsx"
wb_partners.save(partners_file)
print(f"Created {partners_file}")

print(f"\nAll files saved to: {OUTPUT_DIR}")
print("Ready to import!")
