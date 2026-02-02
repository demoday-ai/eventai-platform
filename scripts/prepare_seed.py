"""Prepare seed data from checkpoint forms and past demo day projects.

Extracts projects from data/test/checkpoint12_anon.xlsx and joins
tags from docs/00-research/past-demoday-projects.md.

Output: data/seed/projects_seed.json
"""

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
CHECKPOINT_FILE = ROOT / "data" / "test" / "checkpoint12_anon.xlsx"
PAST_PROJECTS_FILE = ROOT / "docs" / "00-research" / "past-demoday-projects.md"
OUTPUT_FILE = ROOT / "data" / "seed" / "projects_seed.json"

# Sheets with project data (col indices: 2=ФИО, 3=TG, 6=title, 5=track, 12=plan, 13=result)
TRACK_SHEETS = [
    "1КР 13.10",
    "Научный трек",
    "Образовательный трек",
    "Индустриальный трек",
    "Стартап трек",
    "AI Product Альфа",
]


def parse_checkpoint(filepath: Path) -> dict[str, dict]:
    """Extract unique projects from checkpoint xlsx."""
    wb = openpyxl.load_workbook(filepath, read_only=True)
    projects: dict[str, dict] = {}

    for sheet_name in TRACK_SHEETS:
        if sheet_name not in wb.sheetnames:
            print(f"  Warning: sheet '{sheet_name}' not found, skipping")
            continue

        ws = wb[sheet_name]
        headers = None

        for row in ws.iter_rows(values_only=True):
            if headers is None:
                headers = list(row)
                continue

            # Find columns by header name patterns
            title_idx = _find_col(headers, "Название проекта")
            author_idx = _find_col(headers, "ФИО")
            tg_idx = _find_col(headers, "Телеграмм")
            track_idx = _find_col(headers, "Трек")
            plan_idx = _find_col(headers, "План работы")
            result_idx = _find_col(headers, "Ожидаемый результат")

            if title_idx is None:
                # Try positional fallback for known sheets
                if len(row) > 6:
                    title_idx = 6
                    author_idx = 2
                    tg_idx = 3
                    track_idx = 5
                    plan_idx = 12 if len(row) > 12 else None
                    result_idx = 13 if len(row) > 13 else None
                else:
                    continue

            title = _safe_str(row, title_idx).strip()
            if not title or title == "-" or len(title) < 3:
                continue

            # Skip if already seen (dedup by title)
            if title in projects:
                continue

            author = _safe_str(row, author_idx).strip() or "Неизвестен"
            tg = _safe_str(row, tg_idx).strip() or "@unknown"
            track = _safe_str(row, track_idx).strip()

            plan = _safe_str(row, plan_idx).strip() if plan_idx else ""
            result = _safe_str(row, result_idx).strip() if result_idx else ""

            description = ""
            if plan:
                description = plan
            if result:
                description = f"{description}\n\nОжидаемый результат: {result}" if description else result
            if not description:
                description = f"Проект: {title}"
            if track:
                description = f"[Трек: {track}] {description}"

            projects[title] = {
                "title": title,
                "description": description[:2000],  # truncate long descriptions
                "author": author,
                "telegram_contact": tg,
                "tags": [],
            }

    wb.close()
    return projects


def _find_col(headers: list, pattern: str) -> int | None:
    """Find column index by header name pattern."""
    for i, h in enumerate(headers):
        if h and pattern.lower() in str(h).lower():
            return i
    return None


def _safe_str(row: tuple, idx: int | None) -> str:
    """Safely get string value from row."""
    if idx is None or idx >= len(row) or row[idx] is None:
        return ""
    return str(row[idx])


def parse_tags_from_markdown(filepath: Path) -> dict[str, list[str]]:
    """Extract project→tags mapping from past-demoday-projects.md.

    Parses markdown tables with formats:
    - Demo Day: | 10:30 | ProjectName | Tags |
    - Research:  | 1 | ProjectName | Tag |
    """
    tags_map: dict[str, list[str]] = {}
    text = filepath.read_text(encoding="utf-8")

    # Universal pattern: any table row with 3+ columns where first col is
    # time (HH:MM), hall number, or "Зал N"
    # Skip header/separator rows
    row_pattern = re.compile(
        r"^\|\s*([^|]+?)\s*\|\s*([^|]+?)\s*\|\s*([^|]+?)\s*\|",
        re.MULTILINE,
    )

    skip_values = {
        "время", "проект", "теги", "тег", "зал", "тематика", "кол-во",
        "кол-во проектов", "-----", "------", "-------",
    }

    for match in row_pattern.finditer(text):
        col1 = match.group(1).strip()
        col2 = match.group(2).strip()
        col3 = match.group(3).strip()

        # Skip headers and separators
        if col1.lower() in skip_values or col2.lower() in skip_values:
            continue
        if col1.startswith("-") or col2.startswith("-"):
            continue

        # Col1 should be time (HH:MM) or hall number
        is_time = bool(re.match(r"^\d{1,2}:\d{2}$", col1))
        is_hall = bool(re.match(r"^\d+$", col1))
        if not is_time and not is_hall:
            continue

        project_name = col2.strip()
        tags_str = col3.strip()

        if not project_name or not tags_str:
            continue

        tags = [t.strip() for t in tags_str.split(",") if t.strip()]
        if tags and project_name not in tags_map:
            tags_map[project_name] = tags

    return tags_map


def _normalize(s: str) -> str:
    """Normalize title for fuzzy matching: lowercase, strip punctuation/quotes."""
    s = s.lower().strip()
    # Remove common decorators: quotes, dashes, extra spaces
    s = re.sub(r'[«»""\'"\-–—:()]', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def _match_tags(projects: dict[str, dict], tags_map: dict[str, list[str]]) -> int:
    """Match tags to projects using exact + fuzzy matching."""
    matched = 0

    # Build normalized lookup from markdown tags
    norm_tags: dict[str, list[str]] = {}
    for md_title, tags in tags_map.items():
        norm_tags[_normalize(md_title)] = tags

    for title, proj in projects.items():
        # 1. Exact match
        if title in tags_map:
            proj["tags"] = tags_map[title]
            matched += 1
            continue

        # 2. Normalized match
        norm_title = _normalize(title)
        if norm_title in norm_tags:
            proj["tags"] = norm_tags[norm_title]
            matched += 1
            continue

        # 3. Substring match: markdown title contained in checkpoint title or vice versa
        for md_norm, tags in norm_tags.items():
            if len(md_norm) >= 5 and (md_norm in norm_title or norm_title in md_norm):
                proj["tags"] = tags
                matched += 1
                break

    return matched


def main():
    print(f"Reading checkpoint data from {CHECKPOINT_FILE}...")
    projects = parse_checkpoint(CHECKPOINT_FILE)
    print(f"  Found {len(projects)} unique projects")

    print(f"Reading tags from {PAST_PROJECTS_FILE}...")
    tags_map = parse_tags_from_markdown(PAST_PROJECTS_FILE)
    print(f"  Found tags for {len(tags_map)} projects in markdown")

    # Join tags with fuzzy matching
    matched = _match_tags(projects, tags_map)
    print(f"  Matched tags for {matched}/{len(projects)} projects")

    # Convert to list
    result = list(projects.values())

    # Write output
    OUTPUT_FILE.parent.mkdir(parents=True, exist_ok=True)
    with open(OUTPUT_FILE, "w", encoding="utf-8") as f:
        json.dump(result, f, ensure_ascii=False, indent=2)

    print(f"Wrote {len(result)} projects to {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
