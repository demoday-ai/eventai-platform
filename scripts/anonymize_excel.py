"""
Anonymize 4 Demo Day Excel files for testing and demo.

Replaces personal data (FIO, contacts, companies) with consistent anonymous IDs.
Preserves: project names, tracks, scores, criteria, dates, structure.

Usage: python scripts/anonymize_excel.py
Output: data/test/*.xlsx
"""

import json
import re
from pathlib import Path
from collections import defaultdict
import openpyxl

# Paths
DOWNLOADS = Path(r"C:\Users\Admin\Downloads")
REPO_ROOT = Path(r"C:\Users\Admin\Desktop\AI Talent Camp\demoday-core")
OUTPUT_DIR = REPO_ROOT / "data" / "test"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

INPUT_FILES = {
    "experts_23jan": DOWNLOADS / "таблицы с оценками от экспертов DD 2026 23 января.xlsx",
    "experts_22jan": DOWNLOADS / "таблицы с оценками от экспертов DD 2026 22 января.xlsx",
    "checkpoint3": DOWNLOADS / "Ответы из формы 3 контрольного рубежа.xlsx",
    "checkpoint12": DOWNLOADS / "Ответы из форм 1 и 2 контрольных рубежей.xlsx",
}

# --- Name mappings ---
class Anonymizer:
    def __init__(self):
        self.student_map = {}  # FIO -> "Студент_NNN"
        self.expert_map = {}   # name -> "Эксперт_NNN"
        self.mentor_map = {}   # name -> "Ментор_NNN"
        self.company_map = {}  # company -> "Компания_NNN"
        self.telegram_map = {} # @handle -> "@user_NNN"
        self.email_map = {}    # email -> "user_NNN@example.com"
        self.counters = defaultdict(int)

        # Load existing expert mapping for consistency
        mapping_file = REPO_ROOT / "data" / "expert-mapping.json"
        if mapping_file.exists():
            with open(mapping_file, "r", encoding="utf-8") as f:
                mapping = json.load(f)
            for entry in mapping:
                self.expert_map[entry["name"].strip()] = entry["id"]
                if entry.get("telegram"):
                    self.telegram_map[entry["telegram"].strip()] = f"@expert_{entry['id'].split('-')[1]}"

    def _next_id(self, prefix):
        self.counters[prefix] += 1
        return f"{prefix}_{self.counters[prefix]:03d}"

    def anon_student(self, name):
        if not name:
            return name
        name = str(name).strip()
        if not name or name.lower() == "none":
            return name
        if name not in self.student_map:
            self.student_map[name] = self._next_id("Студент")
        return self.student_map[name]

    def anon_expert(self, name):
        if not name:
            return name
        name = str(name).strip()
        if not name or name.lower() == "none":
            return name
        # Already anonymous placeholders
        if re.match(r"^Эксперт\s*\d+$", name):
            return name
        if name not in self.expert_map:
            self.expert_map[name] = self._next_id("Эксперт")
        return self.expert_map[name]

    def anon_mentor(self, name):
        if not name:
            return name
        name = str(name).strip()
        if not name or name.lower() == "none":
            return name
        if name not in self.mentor_map:
            self.mentor_map[name] = self._next_id("Ментор")
        return self.mentor_map[name]

    def anon_telegram(self, handle):
        if not handle:
            return handle
        handle = str(handle).strip()
        if not handle or handle.lower() == "none":
            return handle
        if handle not in self.telegram_map:
            self.telegram_map[handle] = self._next_id("@user")
        return self.telegram_map[handle]

    def anon_email(self, email):
        if not email:
            return email
        email = str(email).strip()
        if not email or email.lower() == "none":
            return email
        if email not in self.email_map:
            n = self._next_id("user")
            self.email_map[email] = f"{n}@example.com"
        return self.email_map[email]

    def anon_company(self, company):
        if not company:
            return company
        company = str(company).strip()
        if not company or company.lower() == "none":
            return company
        if company not in self.company_map:
            self.company_map[company] = self._next_id("Компания")
        return self.company_map[company]

    def anon_contact(self, value):
        """Anonymize a contact field that may contain telegram, email, phone, or name."""
        if not value:
            return value
        val = str(value).strip()
        if not val or val.lower() == "none":
            return val
        # Telegram handle
        if val.startswith("@"):
            return self.anon_telegram(val)
        # Email
        if "@" in val and "." in val:
            return self.anon_email(val)
        # Phone
        if re.match(r"^[\d\s\+\-\(\)]{7,}$", val):
            return "+7-XXX-XXX-XXXX"
        # Likely a name with contact info mixed
        return self.anon_mentor(val)

    def anon_fio_contact_field(self, value):
        """Anonymize fields like 'ФИО и контакт руководителя/ментора'."""
        if not value:
            return value
        val = str(value).strip()
        if not val or val.lower() == "none":
            return val
        return self.anon_mentor(val)

    def anon_free_text_names(self, text, known_names=None):
        """Remove names from free-text comments. Light approach - just return as-is
        since comments rarely contain structured PII beyond what's in named fields."""
        return text


def _is_cyrillic_name(text):
    """Check if text looks like a Cyrillic person name (2-3 words, all Cyrillic)."""
    if not text:
        return False
    words = text.strip().split()
    if len(words) < 2 or len(words) > 4:
        return False
    return all(re.match(r'^[А-ЯЁа-яё\-]+$', w) for w in words)


def anonymize_expert_file(anon, input_path, output_path):
    """Anonymize expert evaluation file (22 or 23 Jan).

    Structure per sheet (variable block size):
    - Row 0: Room name + topic
    - Row 2: Comment for experts
    - Then repeating blocks:
      - Time + "Артефакты"
      - Headers (ФИО, Название проекта, criteria...)
      - Student FIO + Project name + criteria descriptions
      - N expert rows (variable: 6 in 23 Jan, up to 8+ in 22 Jan)
      - Empty row

    Edge cases:
    - Some sheets have cols A/B swapped (project name in A, FIO in B)
    - Standalone comment rows after blocks (expert name + text)
    """
    wb = openpyxl.load_workbook(input_path)

    for ws in wb.worksheets:
        rows_data = []
        for row in ws.iter_rows():
            rows_data.append(row)

        # Find blocks: look for rows where col A contains "ФИО"
        i = 0
        while i < len(rows_data):
            row = rows_data[i]
            cell_a = row[0].value

            if cell_a and "ФИО" in str(cell_a):
                # This is a header row. Next row is student + project.
                if i + 1 < len(rows_data):
                    student_row = rows_data[i + 1]
                    col_a_val = str(student_row[0].value).strip() if student_row[0].value else ""
                    col_b_val = str(student_row[1].value).strip() if len(student_row) > 1 and student_row[1].value else ""

                    # Detect swapped columns: if col B looks like FIO and col A doesn't
                    if col_b_val and _is_cyrillic_name(col_b_val.split("\n")[0]):
                        # Col B has FIO — anonymize it, keep col A (project name)
                        if "\n" in col_b_val:
                            parts = col_b_val.split("\n")
                            anon_parts = [anon.anon_student(p.strip()) if p.strip() and _is_cyrillic_name(p.strip()) else p for p in parts]
                            student_row[1].value = "\n".join(anon_parts)
                        else:
                            student_row[1].value = anon.anon_student(col_b_val)
                    # Normal case: col A has FIO
                    if col_a_val and col_a_val != "None":
                        if "\n" in col_a_val:
                            parts = col_a_val.split("\n")
                            anon_parts = [anon.anon_student(p.strip()) if p.strip() and "ФИО" not in p else p for p in parts]
                            student_row[0].value = "\n".join(anon_parts)
                        elif _is_cyrillic_name(col_a_val):
                            student_row[0].value = anon.anon_student(col_a_val)
                        elif ", " in col_a_val:
                            # Multiple FIOs separated by comma
                            student_row[0].value = anon.anon_student(col_a_val)

                # Expert rows: scan from i+2 until empty row or next block marker
                j = i + 2
                while j < len(rows_data):
                    expert_row = rows_data[j]
                    cell_val = expert_row[0].value
                    if not cell_val:
                        break
                    name = str(cell_val).strip()
                    if not name or name == "None":
                        break
                    if "ФИО" in name or "Артефакты" in name or "Комментарий" in name:
                        break
                    if re.match(r"^\d{1,2}:\d{2}", name):
                        break
                    expert_row[0].value = anon.anon_expert(name)
                    j += 1

                # After block: check for standalone comment rows (name + comment, no empty row before)
                # These appear after the empty row as "Name" + "Comment text" in col B
                k = j + 1
                while k < len(rows_data):
                    post_row = rows_data[k]
                    post_val = post_row[0].value
                    if not post_val:
                        break
                    post_name = str(post_val).strip()
                    if not post_name or post_name == "None":
                        break
                    if re.match(r"^\d{1,2}:\d{2}", post_name):
                        break
                    if "ФИО" in post_name or "Артефакты" in post_name:
                        break
                    # Standalone name + comment
                    if _is_cyrillic_name(post_name):
                        post_row[0].value = anon.anon_expert(post_name)
                    k += 1

                i = j + 1
            else:
                i += 1

    wb.save(output_path)
    print(f"  Saved: {output_path.name}")


def anonymize_checkpoint3(anon, input_path, output_path):
    """Anonymize checkpoint 3 responses (80 columns).

    PII columns:
    - 3: Для кого заполняете (may be FIO)
    - 4: Представьтесь, укажите ФИ
    - 5: Как связаться (telegram/email/phone)
    - 11: ФИО и контакт руководителя
    - 44: Город
    - 60: Компания
    - 61: Должность (keep - not PII per se, but remove specific ones)
    - 62: Должность (свободная)
    - 63: Интересные компании
    - 64: Нежелательные компании
    """
    wb = openpyxl.load_workbook(input_path)

    PII_COLS = {
        3: "student",      # Для кого заполняете форму (FIO)
        4: "student",      # Представьтесь, ФИ
        5: "contact",      # Контакт
        11: "mentor",      # ФИО и контакт руководителя
        41: "freetext_name",  # Где магистрант принес наибольший эффект (may have FIO)
        43: "freetext_name",  # Пожелания/напутствия (may have FIO)
        44: "city",        # Город -> keep (not critical PII for demo)
        60: "company",     # Компания
        63: "company",     # Интересные компании
        64: "company",     # Нежелательные компании
    }

    for ws in wb.worksheets:
        for row_idx, row in enumerate(ws.iter_rows()):
            if row_idx == 0:  # Skip header
                continue
            for col_idx, cell in enumerate(row):
                if col_idx not in PII_COLS or not cell.value:
                    continue
                pii_type = PII_COLS[col_idx]
                val = str(cell.value).strip()
                if not val or val.lower() == "none":
                    continue

                if pii_type == "student":
                    cell.value = anon.anon_student(val)
                elif pii_type == "contact":
                    cell.value = anon.anon_contact(val)
                elif pii_type == "mentor":
                    cell.value = anon.anon_fio_contact_field(val)
                elif pii_type == "company":
                    cell.value = anon.anon_company(val)
                elif pii_type == "freetext_name":
                    # Only anonymize if the entire cell looks like a name
                    if _is_cyrillic_name(val):
                        cell.value = anon.anon_student(val)

    # Handle sheets without headers (e.g., "Потеряшки до 17")
    # These have same column structure but no header row
    for ws in wb.worksheets:
        # Check if first row looks like data (no header keywords)
        first_row = list(ws.iter_rows(min_row=1, max_row=1))[0]
        has_header = any(
            c.value and ("ID" == str(c.value).strip() or "Время" in str(c.value) or "Выберите" in str(c.value))
            for c in first_row
        )
        if has_header:
            continue  # Already processed above

        # No header — anonymize FIO columns (3, 4) and contact (5) by position
        for row in ws.iter_rows():
            for col_idx, cell in enumerate(row):
                if not cell.value:
                    continue
                val = str(cell.value).strip()
                if not val or val == "None":
                    continue
                if col_idx == 3 and _is_cyrillic_name(val):
                    cell.value = anon.anon_student(val)
                elif col_idx == 4 and _is_cyrillic_name(val):
                    cell.value = anon.anon_student(val)
                elif col_idx == 5:
                    cell.value = anon.anon_contact(val)

    wb.save(output_path)
    print(f"  Saved: {output_path.name}")


def anonymize_checkpoint12(anon, input_path, output_path):
    """Anonymize checkpoints 1 & 2 responses.

    Main sheet '1КР 13.10' columns:
    - 2: ФИО
    - 3: Телеграмм
    - 14: Команда проекта (list of names)

    Other sheets have similar structure.
    PII columns by index (0-based):
    - 2: ФИО
    - 3: Телеграмм (для связи)
    - 14: Команда проекта (free text with names)
    """
    wb = openpyxl.load_workbook(input_path)

    for ws in wb.worksheets:
        # Detect header row to find PII columns dynamically
        header_row = None
        for row_idx, row in enumerate(ws.iter_rows(max_row=3)):
            for cell in row:
                if cell.value and "ФИО" in str(cell.value):
                    header_row = row_idx
                    break
            if header_row is not None:
                break

        if header_row is None:
            # Sheet without standard headers (e.g., Потеряшки) - anonymize by position
            # These sheets may have data without headers
            for row_idx, row in enumerate(ws.iter_rows()):
                for col_idx, cell in enumerate(row):
                    if not cell.value:
                        continue
                    val = str(cell.value).strip()
                    # Detect telegram handles anywhere
                    if val.startswith("@") and len(val) > 2:
                        cell.value = anon.anon_telegram(val)
                    # Detect emails
                    elif "@" in val and "." in val and " " not in val and len(val) < 60:
                        cell.value = anon.anon_email(val)
            continue

        # Find PII column indices from headers
        pii_indices = {}
        headers = list(ws.iter_rows(min_row=header_row+1, max_row=header_row+1))[0]
        for col_idx, cell in enumerate(headers):
            if not cell.value:
                continue
            h = str(cell.value).strip().lower()
            if "фио" in h and "руковод" not in h and "ментор" not in h:
                pii_indices[col_idx] = "student"
            elif "телеграм" in h or "telegram" in h:
                pii_indices[col_idx] = "telegram"
            elif "команда" in h:
                pii_indices[col_idx] = "team"
            elif ("фио" in h or "руковод" in h or "ментор" in h) and ("контакт" in h or "руковод" in h):
                pii_indices[col_idx] = "mentor"
            elif "руковод" in h or "ментор" in h:
                pii_indices[col_idx] = "mentor"

        # Anonymize data rows
        for row_idx, row in enumerate(ws.iter_rows()):
            if row_idx <= header_row:
                continue
            for col_idx, cell in enumerate(row):
                if col_idx not in pii_indices or not cell.value:
                    continue
                val = str(cell.value).strip()
                if not val or val.lower() == "none":
                    continue

                pii_type = pii_indices[col_idx]
                if pii_type == "student":
                    cell.value = anon.anon_student(val)
                elif pii_type == "telegram":
                    cell.value = anon.anon_telegram(val)
                elif pii_type == "mentor":
                    cell.value = anon.anon_fio_contact_field(val)
                elif pii_type == "team":
                    # Team field may contain multiple names separated by commas/newlines
                    names = re.split(r"[,;\n]", val)
                    anonymized = [anon.anon_student(n.strip()) for n in names if n.strip()]
                    cell.value = ", ".join(anonymized)

    wb.save(output_path)
    print(f"  Saved: {output_path.name}")


def sweep_all_cells(anon, file_path):
    """Final sweep: scan ALL cells for any remaining PII patterns.

    Catches:
    - Telegram handles (@username) embedded in free text
    - Emails embedded in free text
    - Cyrillic full names (Фамилия Имя [Отчество]) embedded in text
    - FIO with trailing punctuation (quotes, etc.)
    - Multiple names separated by commas or "и"
    - Does NOT touch URLs (google drive, medium, etc.)
    """
    wb = openpyxl.load_workbook(file_path)
    fixes = 0

    # Regex for Cyrillic full name (2-3 words, each starting with capital)
    cyrillic_name_re = re.compile(
        r'(?<![А-ЯЁа-яё])([А-ЯЁ][а-яё]{1,20}\s+[А-ЯЁ][а-яё]{1,20}(?:\s+[А-ЯЁ][а-яё]{1,20})?)(?![А-ЯЁа-яё])'
    )

    # Known non-name patterns to skip
    skip_patterns = {
        "Студент_", "Эксперт_", "Ментор_", "Компания_",
        "Название проекта", "Название комнаты", "Комментарий для",
        "Практическая значимость", "Оценка импакта", "Потенциал масштабирования",
        "Технологическая сложность", "Качество реализации", "Свободные комментарии",
        "Итоговый процент", "Talent Matcher", "Demo Day", "Talent Hub",
        "Яндекс Маркет", "Норильский Никель",
    }

    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if not cell.value:
                    continue
                val = str(cell.value)

                # Skip cells that are clearly already anonymized or structural
                if any(skip in val for skip in skip_patterns):
                    # Still check for telegram/email
                    pass
                else:
                    # Check for Cyrillic names
                    name_matches = cyrillic_name_re.findall(val)
                    for name_match in name_matches:
                        nm = name_match.strip()
                        # Skip common non-name phrases
                        if any(w in nm.lower() for w in [
                            "контрольн", "научн", "индустриальн", "образовательн",
                            "стартап", "курс", "трек", "проект", "семестр",
                            "степень вовлечённости", "какой области", "продолжи",
                        ]):
                            continue
                        # Must have all words starting with uppercase
                        words = nm.split()
                        if not all(w[0].isupper() for w in words):
                            continue
                        # Skip short common words
                        if len(words) == 2 and any(len(w) <= 2 for w in words):
                            continue
                        # This is likely a name — anonymize in context
                        anon_name = anon.anon_student(nm)
                        val = val.replace(nm, anon_name)
                        fixes += 1

                # Find embedded telegram handles (not in URLs)
                tg_matches = re.findall(r'(?<!\w)(@[a-zA-Z_]\w{2,})', val)
                for tg in tg_matches:
                    if "@user_" in tg or "@expert_" in tg:
                        continue
                    idx = val.find(tg)
                    if idx > 0 and val[idx-1] == '/':
                        continue
                    anon_tg = anon.anon_telegram(tg)
                    val = val.replace(tg, anon_tg)
                    fixes += 1

                # Find embedded emails
                email_matches = re.findall(r'[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}', val)
                for em in email_matches:
                    if "@example.com" in em:
                        continue
                    if any(svc in em for svc in ["google.com", "drive.google", "medium.com", "github.com"]):
                        continue
                    anon_em = anon.anon_email(em)
                    val = val.replace(em, anon_em)
                    fixes += 1

                if str(cell.value) != val:
                    cell.value = val

    wb.save(file_path)
    return fixes


def main():
    anon = Anonymizer()

    print("Anonymizing expert evaluations (23 Jan)...")
    anonymize_expert_file(
        anon,
        INPUT_FILES["experts_23jan"],
        OUTPUT_DIR / "experts_23jan_anon.xlsx",
    )

    print("Anonymizing expert evaluations (22 Jan)...")
    anonymize_expert_file(
        anon,
        INPUT_FILES["experts_22jan"],
        OUTPUT_DIR / "experts_22jan_anon.xlsx",
    )

    print("Anonymizing checkpoint 3...")
    anonymize_checkpoint3(
        anon,
        INPUT_FILES["checkpoint3"],
        OUTPUT_DIR / "checkpoint3_anon.xlsx",
    )

    print("Anonymizing checkpoints 1 & 2...")
    anonymize_checkpoint12(
        anon,
        INPUT_FILES["checkpoint12"],
        OUTPUT_DIR / "checkpoint12_anon.xlsx",
    )

    # Final sweep: catch any remaining PII in all files
    print("\nFinal sweep for remaining PII...")
    total_fixes = 0
    for out_file in OUTPUT_DIR.glob("*_anon.xlsx"):
        fixes = sweep_all_cells(anon, out_file)
        if fixes:
            print(f"  {out_file.name}: {fixes} additional fixes")
        total_fixes += fixes
    print(f"  Total sweep fixes: {total_fixes}")

    # Save mapping for reference (useful for debugging, NOT for commit)
    mapping = {
        "students": anon.student_map,
        "experts": {k: v for k, v in anon.expert_map.items()},
        "mentors": anon.mentor_map,
        "companies": anon.company_map,
        "telegrams": anon.telegram_map,
    }
    mapping_path = OUTPUT_DIR / "_anonymization_mapping.json"
    with open(mapping_path, "w", encoding="utf-8") as f:
        json.dump(mapping, f, ensure_ascii=False, indent=2)

    print(f"\nMapping saved to: {mapping_path.name}")
    print("\nStats:")
    print(f"  Students:  {len(anon.student_map)}")
    print(f"  Experts:   {len(anon.expert_map)}")
    print(f"  Mentors:   {len(anon.mentor_map)}")
    print(f"  Companies: {len(anon.company_map)}")
    print(f"  Telegrams: {len(anon.telegram_map)}")
    print(f"\nAll anonymized files saved to: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
